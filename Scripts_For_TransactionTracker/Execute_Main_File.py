from selenium import webdriver
from openpyxl import *
import time
from Utilites import AppConstants
from Scripts_For_TransactionTracker.PurchaseOrder_from_TT import Po_from_TT
from Scripts_For_TransactionTracker.FTP_Operations import FTP_Operation
from Scripts_For_TransactionTracker.validation_Script import Edi_Validator
from AppResources import ElementLocator_For_TransactionTracker as ElementLocator
from Utilites.LogFileUtility import LogFileUtility as lo

class Execute_Main:

    def __init__(self):
        try:
            print("In Execute_Main Constructor")
            chrome_options = webdriver.ChromeOptions()
            # chrome_options.add_argument("--incognito")
            chrome_options.add_argument("--start-maximized")
            self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER, chrome_options=chrome_options)
            self.log = lo
        except:
            self.log.log_to_file(self, "ERROR", "NOT able to Connected Scripts.Execute_Main() Constructor")

    def Execute_PurchaseOrder_from_TT(self):
        try:
            self.log.log_to_file(self, "INFO", " Executing Execurte_Main_FIle.Execute_PurchaseOrder_from_TT()")
            file_Po = Po_from_TT(self.log, self.v_driver)
            file_Po.login("Launchpad")
            v_input_wb = load_workbook(ElementLocator.INPUT_FILE_PATH)
            v_input_sheet = v_input_wb.get_sheet_by_name("TT_Input")
            for i in range(2, v_input_sheet.max_row + 1):
                Supplier_Name = str(v_input_sheet.cell(row=i, column=1).value)
                Receiver_Name = str(v_input_sheet.cell(row=i, column=2).value)
                Doc_type = str(v_input_sheet.cell(row=i, column=3).value)
                Date = str(v_input_sheet.cell(row=i, column=4).value)
                print(Supplier_Name,Receiver_Name,Doc_type,Date)
                file_Po.TT_Operations(Supplier_Name,Receiver_Name,Doc_type,Date)
                break
        except Exception as e:
            self.log.log_to_file(self,"ERROR",f"Exception is {str(e)} in Execute_Maine.Execute_PurchaseOrder_from_TT()")

    def Execute_Validator(self):
        try:
            self.log.log_to_file(self, "INFO", " Executing Execure_Main_FIle.Execute_Validation()")
            v_input_wb = load_workbook(ElementLocator.INPUT_FILE_PATH)
            v_input_sheet = v_input_wb.get_sheet_by_name("Input")
            for i in range(2, v_input_sheet.max_row + 1):
                Sender_Qualifiers = v_input_sheet.cell(row=i, column=1).value
                Receiver_Qualifiers = v_input_sheet.cell(row=i, column=2).value
                FileLocation = v_input_sheet.cell(row=i, column=3).value
                Supplier_Version = str(v_input_sheet.cell(row=i, column=4).value)
                Retailer_Version = str(v_input_sheet.cell(row=i, column=5).value)
                GS_Supplier_Version = str(v_input_sheet.cell(row=i, column=6).value)
                GS_Retailer_Version = str(v_input_sheet.cell(row=i, column=7).value)
                Document_version = str(v_input_sheet.cell(row=i, column=8).value)
                file_validation = Edi_Validator(self.log,Sender_Qualifiers,Receiver_Qualifiers,FileLocation,
                                                Supplier_Version,Retailer_Version,GS_Supplier_Version,
                                                GS_Retailer_Version,Document_version)
                file_validation.remove_delimiters()
                file_validation.presence_of_U()
                file_validation.validate_supplier_retailer_cability_version()
                file_validation.ST_and_GE_Validation()
                file_validation.ST_and_SE_Validation()
                file_validation.Interchange_Header_Segment_Validation()
                file_validation.spacing_validation()
                file_validation.Document_versions()
                file_validation.validation_for_Qualifiers()
        except Exception as e:
            self.log.log_to_file(self, "ERROR", f"Exception is {str(e)} in Execute_Maine().Execute_Validation()")

    def Execute_FTP(self):
        file_Po = Po_from_TT(self.log, self.v_driver)
        # Ftp_File = FTP_Operation(self.log, self.v_driver)
        # self.v_driver.execute_script("window.open('about:blank', 'tab2');")
        # self.v_driver.switch_to.window("tab2")
        # file_Po.login("FTP Pre-prod")
        # Ftp_File.File_Uploading()
        # time.sleep(3)
        v_input_wb = load_workbook(ElementLocator.INPUT_FILE_PATH)
        v_input_sheet = v_input_wb.get_sheet_by_name("TT_Input")
        for i in range(2, v_input_sheet.max_row + 1):
            Supplier_Name = str(v_input_sheet.cell(row=i, column=1).value)
            Receiver_Name = str(v_input_sheet.cell(row=i, column=2).value)
            Doc_type = str(v_input_sheet.cell(row=i, column=3).value)
            Date = str(v_input_sheet.cell(row=i, column=4).value)
            print(Supplier_Name, Receiver_Name, Doc_type, Date)
            # self.v_driver.switch_to.window(self.v_driver.window_handles[0])
            # time.sleep(3)
            file_Po.call_check_accepted(Supplier_Name,Receiver_Name,Doc_type,Date)
            break

